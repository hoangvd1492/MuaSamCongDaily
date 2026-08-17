from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from src.helpers import (
    map_quy_trinh_ap_dung,
    map_hinh_thuc_lua_chon_nha_thau,
    map_phuong_thuc_lua_chon_nha_thau,
    map_linh_vuc,
    map_loai_hop_dong,
    map_trang_thai_tbmt,
    build_url_view_detail_tbmt,
    build_url_view_detail_khlcnt,
)


def export_tbmt_excel(rows: list, from_time=None, to_time=None) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    sheet_open = wb.create_sheet("TBMT_ChuaDongThau")
    sheet_closed = wb.create_sheet("TBMT_DaDongThau")

    # Danh sách cột & độ rộng (Đã tăng nhẹ độ rộng để vừa size chữ 13-14)
    columns = [
        ("Mã TBMT", 28),
        ("Tên gói thầu", 65),
        ("Giá gói thầu", 24),
        ("Chủ đầu tư", 50),
        ("Thời điểm đóng thầu", 28),
        ("Thời gian đăng tải gốc", 28),
        ("Thời gian sửa TBMT", 28),
        ("Mã KHLCNT", 24),
        ("Tên dự toán mua sắm", 50),
        ("Dự toán mua sắm", 24),
        ("Quy trình áp dụng", 28),
        ("Hình thức LCNT", 28),
        ("Phương thức LCNT", 28),
        ("Lĩnh vực", 18),
        ("Loại hợp đồng", 24),
        ("Thời gian thực hiện gói thầu", 28),
        ("Hiệu lực HSDT", 24),
        ("Số tiền bảo đảm dự thầu", 26),
        ("Thời điểm mở thầu", 28),
        ("Trạng thái TBMT", 28),
        ("Người trúng thầu", 40),
        ("Giá trúng thầu", 24),
        ("Chi tiết TBMT", 20),
        ("Chi tiết KHLCNT", 20),
    ]

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="4472C4")
    highlight_fill = PatternFill("solid", fgColor="C6EFCE")

    def get_val(item, key):
        if item is None:
            return None
        if isinstance(item, dict) or hasattr(item, "keys"):
            return item.get(key)
        return getattr(item, key, None)

    def to_num(val):
        if val is None or val == "":
            return None
        try:
            return float(str(val).replace(",", "").strip())
        except (ValueError, TypeError):
            return None

    def format_dt(val):
        """Chuyển đổi 2026-08-15T18:03:35.334 -> 2026-08-15 18:03:35.334"""
        if not val:
            return None
        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%d %H:%M:%S")
        val_str = str(val).replace("T", " ").replace("Z", "").strip()
        if "+" in val_str:
            val_str = val_str.split("+")[0].strip()
        return val_str

    for ws in [sheet_open, sheet_closed]:
        # Dòng 1: Tiêu đề báo cáo (Font Size 18)
        ws.merge_cells("A1:X1")
        ws["A1"] = f"THỜI ĐIỂM LẤY DỮ LIỆU: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A1"].font = Font(name="Calibri", size=18, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 50

        # Dòng 2: Header cột (Font Size 14)
        ws.append([c[0] for c in columns])
        ws.row_dimensions[2].height = 40

        for col_idx, (_, width) in enumerate(columns, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width
            cell = ws.cell(row=2, column=col_idx)
            cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.freeze_panes = "A3"

    for row in rows:
        thoi_diem_dong_thau = format_dt(get_val(row, "thoiDiemDongThau"))
        ngay_dang_tai_goc = format_dt(get_val(row, "ngayDangTaiGoc"))
        thoi_gian_sua = format_dt(get_val(row, "thoiGianSuaTBMT"))
        thoi_diem_mo_thau = format_dt(get_val(row, "thoiDiemMoThau"))

        str_to_time = format_dt(to_time)
        str_from_time = format_dt(from_time)

        # Chọn sheet
        ws = sheet_closed if (str_to_time and thoi_diem_dong_thau and thoi_diem_dong_thau <= str_to_time) else sheet_open

        row_data = [
            get_val(row, "maTBMT"),
            get_val(row, "tenGoiThau"),
            to_num(get_val(row, "giaGoiThau")),
            get_val(row, "chuDauTu"),
            thoi_diem_dong_thau,
            ngay_dang_tai_goc,
            thoi_gian_sua,
            get_val(row, "maKHLCNT"),
            get_val(row, "tenDuToanMuaSam"),
            to_num(get_val(row, "duToanMuaSam")),
            map_quy_trinh_ap_dung(get_val(row, "quyTrinhApDung")),
            map_hinh_thuc_lua_chon_nha_thau(get_val(row, "hinhThucLuaChonNhaThau")),
            map_phuong_thuc_lua_chon_nha_thau(get_val(row, "phuongThucLuaChonNhaThau")),
            map_linh_vuc(get_val(row, "linhVuc")),
            map_loai_hop_dong(get_val(row, "loaiHopDong")),
            get_val(row, "thoiGianThucHienGoiThau"),
            get_val(row, "hieuLucHoSoDuThau"),
            to_num(get_val(row, "soTienBaoDamDuThau")),
            thoi_diem_mo_thau,
            map_trang_thai_tbmt(get_val(row, "trangThaiTBMT")),
            get_val(row, "nguoiTrungThau"),
            to_num(get_val(row, "giaTrungThau")),
        ]

        ws.append(row_data)
        curr_row = ws.max_row
    

        # Hyperlink TBMT
        link_tbmt = build_url_view_detail_tbmt(get_val(row, "id"), get_val(row, "quyTrinhApDung"))
        cell_tbmt = ws.cell(row=curr_row, column=23, value="Xem chi tiết")
        if link_tbmt:
            cell_tbmt.hyperlink = link_tbmt

        # Hyperlink KHLCNT
        link_kh = build_url_view_detail_khlcnt(get_val(row, "planId"), get_val(row, "maKHLCNT"))
        cell_kh = ws.cell(row=curr_row, column=24, value="Xem chi tiết")
        if link_kh:
            cell_kh.hyperlink = link_kh

        # Highlight nếu sửa trong khoảng crawl
        if str_from_time and str_to_time and thoi_gian_sua:
            if str_from_time < thoi_gian_sua <= str_to_time:
                for cell in ws[curr_row]:
                    cell.fill = highlight_fill

    # Format toàn bộ dòng dữ liệu (Font Size 13)
    for ws in [sheet_open, sheet_closed]:
        if ws.max_row < 3:
            continue

        for r_idx in range(3, ws.max_row + 1):
            for c_idx in range(1, 25):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.border = border
                cell.font = Font(name="Calibri", size=13)
                cell.alignment = Alignment(vertical="center", wrap_text=True)

            # Cột số tiền: Format số và căn phải
            for col_l in ["C", "J", "R", "V"]:
                c = ws[f"{col_l}{r_idx}"]
                if c.value is not None:
                    c.number_format = "#,##0"
                c.alignment = Alignment(horizontal="right", vertical="center")

            # Cột ngày giờ: Căn giữa
            for col_l in ["E", "F", "G", "S"]:
                ws[f"{col_l}{r_idx}"].alignment = Alignment(horizontal="center", vertical="center")

            # Cột hyperlink: Màu xanh, gạch chân
            for col_l in ["W", "X"]:
                c = ws[f"{col_l}{r_idx}"]
                c.font = Font(name="Calibri", size=13, color="0563C1", underline="single")
                c.alignment = Alignment(horizontal="center", vertical="center")

        ws.auto_filter.ref = f"A2:X{ws.max_row}"

    output_dir = Path("crawl_data")
    output_dir.mkdir(parents=True, exist_ok=True)
    file_path = output_dir / "TBMT.xlsx"
    wb.save(file_path)

    return str(file_path)